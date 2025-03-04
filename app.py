import streamlit as st
import pandas as pd
from openpyxl import load_workbook, Workbook
from datetime import datetime
import os
import unicodedata  # 文字正規化用

# Excelファイルのパス
DATA_FILE = "waste_log.xlsx"
PRODUCTS_FILE = "products.xlsx"

# 商品データを読み込む
@st.cache_data
def load_data():
    try:
        return pd.read_excel(PRODUCTS_FILE, engine="openpyxl")
    except FileNotFoundError:
        st.error(f"商品データ {PRODUCTS_FILE} が見つかりません")
        return pd.DataFrame(columns=["商品名"])  # 空のDataFrameを返す

df = load_data()

st.title("🐟 fishパンロス記録")

# 🔍 頭文字入力（ひらがな・カタカナ・英字対応）
input_text = st.text_input("商品の頭文字を入力してください", value="", key="input_text").strip()

# 文字正規化（ひらがな⇔カタカナ対応）
def normalize_text(text):
    text = unicodedata.normalize("NFKC", text).lower()  # 全角→半角、大文字→小文字変換
    text = text.translate(str.maketrans("アイウエオカキクケコサシスセソタチツテトナニヌネノ"
                                        "ハヒフヘホマミムメモヤユヨラリルレロワヲン",
                                        "あいうえおかきくけこさしすせそたちつてとなにぬねの"
                                        "はひふへほまみむめもやゆよらりるれろわをん"))
    return text

if input_text:
    normalized_input = normalize_text(input_text)
    
    # 🔍 部分一致検索に変更
    filtered_df = df[df["商品名"].apply(lambda x: normalized_input in normalize_text(str(x)))]

    if not filtered_df.empty:
        selected_product = st.selectbox("商品を選択してください", filtered_df["商品名"])
    else:
        st.warning("該当する商品が見つかりません")
        selected_product = None
else:
    selected_product = None

if selected_product:
    today = datetime.now().strftime("%Y-%m-%d")

    # 📊 既存のデータを読み込んで、本日入力回数を取得
    input_count = 0
    previous_quantity = 0  # すでに登録された数

    if os.path.exists(DATA_FILE):
        book = load_workbook(DATA_FILE)
        sheet = book.active

        # 1行目（日付）を取得
        date_headers = [cell.value for cell in sheet[1] if cell.value]

        if today in date_headers:
            col_index = date_headers.index(today) + 1  # 1ベース
        else:
            col_index = None

        # 商品名のリスト
        product_names = [sheet.cell(row=i, column=1).value for i in range(2, sheet.max_row + 1) if sheet.cell(row=i, column=1).value]

        # すでに記録がある場合の処理
        if selected_product in product_names:
            row_index = product_names.index(selected_product) + 2
            
            # 🔄 **入力回数をカウント**
            input_count = 0
            for i in range(2, sheet.max_row + 1):
                if sheet.cell(row=i, column=1).value == selected_product:
                    if col_index:
                        existing_value = sheet.cell(row=i, column=col_index).value
                        if existing_value:
                            previous_quantity = existing_value  # 既存の数量
                            input_count += 1  # 入力回数をカウント

    # 🔢 入力フォーム（リセット対応）
    if "quantity_input" not in st.session_state:
        st.session_state["quantity_input"] = 0

    quantity = st.number_input(f"{selected_product} の廃棄個数を入力", min_value=0, step=1, value=st.session_state["quantity_input"])

    # 📊 **本日の入力回数を表示**
    st.info(f"📌 本日 {selected_product} の入力回数: {input_count} 回")

    # 🛠 2回目以降の入力処理（合計するか確認）
    if input_count > 0:
        st.info(f"🔄 既に本日 {selected_product} は {previous_quantity} 個 記録されています。")
        confirm = st.checkbox("✅ 既存の数に合計する")
    else:
        confirm = True  # 初回は自動的に記録

    if st.button("記録"):
        try:
            if os.path.exists(DATA_FILE):
                book = load_workbook(DATA_FILE)
                sheet = book.active
            else:
                book = Workbook()
                sheet = book.active
                sheet.cell(row=1, column=1, value="商品名")
                sheet.cell(row=1, column=2, value=today)

            # 既存の日付を取得
            date_headers = [cell.value for cell in sheet[1] if cell.value]

            # 今日の日付がなければ追加
            if today not in date_headers:
                sheet.cell(row=1, column=len(date_headers) + 1, value=today)
                date_headers.append(today)

            # 商品の行を取得
            if selected_product in product_names:
                row_index = product_names.index(selected_product) + 2
            else:
                row_index = sheet.max_row + 1
                sheet.cell(row=row_index, column=1, value=selected_product)

            # 日付の列を取得
            col_index = date_headers.index(today) + 1

            # 📝 2回目以降の入力は合計する
            if confirm:
                new_quantity = previous_quantity + quantity
            else:
                new_quantity = quantity

            sheet.cell(row=row_index, column=col_index, value=new_quantity)

            book.save(DATA_FILE)
            book.close()

            st.success(f"{selected_product} の廃棄 {new_quantity} 個 を記録しました！")

            # 入力欄をリセット
            st.session_state["quantity_input"] = 0

        except Exception as e:
            st.error(f"データの記録に失敗しました: {e}")

# 📊 記録データの表示
if os.path.exists(DATA_FILE):
    st.subheader("📊 廃棄データ")
    waste_df = pd.read_excel(DATA_FILE, engine="openpyxl")
    st.dataframe(waste_df)

# 入力終了ボタン
if st.button("入力を終了する"):
    st.warning("✅ 入力を終了しました。画面を閉じてください。")
    st.stop()
