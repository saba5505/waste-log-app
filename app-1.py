import streamlit as st
import pandas as pd
from openpyxl import load_workbook, Workbook
from datetime import datetime

# Excelファイルのパス
DATA_FILE = "waste_log.xlsx"

# 商品データを読み込む
@st.cache_data
def load_data():
    return pd.read_excel("products.xlsx", engine="openpyxl")

df = load_data()

st.title("fishパンロス")

# 頭文字で商品を検索
input_text = st.text_input("商品の頭文字を入力してください（例: あ）")

if input_text:
    filtered_df = df[df["商品名"].str.startswith(input_text)]
    
    if not filtered_df.empty:
        selected_product = st.selectbox("商品を選択してください", filtered_df["商品名"])
        quantity = st.number_input(f"{selected_product} の廃棄個数を入力", min_value=0, step=1)

        if st.button("記録"):
            # 今日の日付（列として使用）
            today = datetime.now().strftime("%Y-%m-%d")

            try:
                # 既存の Excel を読み込む
                book = load_workbook(DATA_FILE)
                sheet = book.active
                
                # 1行目の日付を取得
                date_headers = [cell.value for cell in sheet[1] if cell.value]

                # 今日の日付がなければ追加
                if today not in date_headers:
                    sheet.cell(row=1, column=len(date_headers) + 2, value=today)  # 日付を追加
                    date_headers.append(today)

                # 商品名のリストを取得
                product_names = [sheet.cell(row=i, column=1).value for i in range(2, sheet.max_row + 1) if sheet.cell(row=i, column=1).value]

                # 商品がなければ追加
                if selected_product not in product_names:
                    new_row = sheet.max_row + 1
                    sheet.cell(row=new_row, column=1, value=selected_product)
                    product_names.append(selected_product)
                else:
                    new_row = product_names.index(selected_product) + 2  # Excel は1ベース

                # 日付の列インデックスを取得
                col_index = date_headers.index(today) + 2
                sheet.cell(row=new_row, column=col_index, value=quantity)  # 廃棄数を入力

                # 保存
                book.save(DATA_FILE)
                book.close()
                st.success(f"{selected_product} の廃棄 {quantity} 個 を記録しました！")

            except FileNotFoundError:
                # ファイルがない場合、新規作成
                book = Workbook()
                sheet = book.active

                # 1行目に日付
                sheet.cell(row=1, column=1, value="商品名")
                sheet.cell(row=1, column=2, value=today)

                # 2行目に商品データ
                sheet.cell(row=2, column=1, value=selected_product)
                sheet.cell(row=2, column=2, value=quantity)

                book.save(DATA_FILE)
                book.close()
                st.success(f"{selected_product} の廃棄 {quantity} 個 を新規ファイルに記録しました！")
    else:
        st.warning("該当する商品が見つかりません")

# 🔴【入力を終了する】ボタン
if st.button("入力を終了する"):
    st.warning("✅ 入力を終了しました。画面を閉じてください。")
    st.stop()  # 以降の処理を止める
